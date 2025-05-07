import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def create_pim_test_data():
    excel_path = os.path.join("TestData", "orangehrm.xlsx")
    
    workbook = load_workbook(excel_path)
    
    if "Sheet2" not in workbook.sheetnames:
        workbook.create_sheet("Sheet2")
    
    sheet = workbook["Sheet2"]
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.value = None
    
    headers = ["First Name", "Middle Name", "Last Name", "Employee ID", "User Created", "User Validated"]
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    test_data = [
        ["Tony", "Edward", "Stark", "IRON001", "", ""],
        ["Thor", "Odinson", "Asgard", "THOR002", "", ""],
        ["Steve", "Grant", "Rogers", "CAP003", "", ""],
        ["Bruce", "Robert", "Banner", "HULK004", "", ""],
        ["Peter", "Benjamin", "Parker", "SPDR005", "", ""],
        ["Natasha", "Alianovna", "Romanoff", "BLKW006", "", ""],
        ["Clint", "Francis", "Barton", "HAWK007", "", ""],
        ["Wanda", "Marya", "Maximoff", "SCWT008", "", ""],
        ["Stephen", "Vincent", "Strange", "DRST009", "", ""],
        ["Carol", "Susan", "Danvers", "CAPM010", "", ""]
    ]
    
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[chr(64 + col)].width = 15
    
    workbook.save(excel_path)
    print("Test data has been added to Sheet2 of the Excel file.")

if __name__ == "__main__":
    create_pim_test_data() 
